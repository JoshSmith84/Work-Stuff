# simple script to fix already created work books to
# highlight devices with missing/disabled TPM
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

folder = 'U:\\Joshua\\Work-Stuff\\AMP\\'
workbook_regex = re.compile(r"""^(.*?)(\.)(xlsx)$""")
filenames = os.listdir(folder)
gold = 'FFD966'

for files in filenames:
    mo = re.search(workbook_regex, files)
    if mo:
        wb_file = folder + mo.group()
        wb = load_workbook(wb_file)
        # sheet = wb['Encryption']
        # max_row = sheet.max_row
        # for i in range(1, max_row + 1):
        #     cell_data = sheet.cell(row=i, column=2).value
        #     if cell_data == 'No TPM Detected':
        #         for cell in sheet[i]:
        #             cell.fill = PatternFill(start_color=gold,
        #                                     end_color=gold,
        #                                     fill_type='solid')
        # board_sheet = wb.create_sheet('On-Offboard')
        # font_header = Font(size=12, bold=True)
        # board_headers = [('Device Name', 'Sophos?', 'Umbrella?',
        #                   'Blackpoint SNAP?', 'Concierge?', 'Arctic Wolf?',
        #                   'AV Defender?', 'Competing AV?')]
        board_sheet = wb['On-Offboard']
        for i in range(1, 9):
            col = get_column_letter(i)
            board_sheet.column_dimensions[col].width = 25
        # for row in board_headers:
        #     board_sheet.append(row)
        # for cell in board_sheet['1:1']:
        #     cell.font = font_header
        # board_sheet.freeze_panes = 'A2'
        wb.save(wb_file)