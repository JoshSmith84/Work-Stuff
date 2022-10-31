#! python3
# Ncentral_AMPs.py - parse AMP output emails from Outlook,
# while also keeping track of client_name
# and job details from output email body. Parse data,
# create/update master client_name excel files, move processed email when done.

# For TPM and BDE: text output only
# For software inventory amp, select "Send task output file in Email"

# One small issue:
# output from devices that reside in sub-sites of a client show that site
# and only that site as the customer.

# To solve this, the job name must be named in ncentral as client name -
# whatever. Will need to specify the same client name exactly including
# case(though I could fold it...) for this to output into same worksheet.
# Regardless, if this method is not followed it will still parse
# client name from body, but that may end up being a site name...

# for now, support for TPM checks, BDE/encryption status, and live "asset scans"

# Author: Josh Smith

import win32com.client
import re
import os
import shelve
from zipfile import ZipFile
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %'
                                                '(levelname)s - %'
                                                '(message)s'
                    )


def device_check(wb_sheet, device: str) -> int:
    """
    Pass the needed sheet name and device name.
    Iterate through the sheet and look for the existence of the device name.
    If the device is present, return the row it resides in,
    if the device is not present return empty value for device.
    Needed modules: openpyxl, Workbook
    :param wb_sheet: The sheet to check
    :param device: The device name to search for
    :return: The row number the device resides in.
    """
    max_row = wb_sheet.max_row
    device_row = ''
    for i in range(1, max_row + 1):
        cell_data = wb_sheet.cell(row=i, column=1).value
        if device in cell_data:
            device_row = i
            break
        else:
            device_row = ''
    return device_row


logging.disable(logging.CRITICAL)
logging.debug('Start of program\n')

# Variable initialization
# Where the Work Stuff shelf resides
db = 'U:\\Joshua\\Dropbox\\Dropbox\\Python\\Work Stuff\\work_stuff'

# Pull email and out folder from shelf
with shelve.open(db) as shelf:
    email = shelf['josh_email']
    parent_f = shelf['out_folder']

# Outlook and folder variables
outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
inbox = outlook.Folders.Item(email).Folders[
    'Inbox'].Folders['Auto Policy']
outbox = outlook.Folders.Item(email).Folders[
    'Inbox'].Folders['Auto Policy'].Folders['Processed']
messages = inbox.Items
err_file = parent_f + 'AMP_errors.txt'

# colors and format variables
gold = 'FFD966'
light_red = 'E06666'
white = 'ffffff'
green = '93c47d'
red_fill = PatternFill(start_color=light_red,
                       end_color=light_red,
                       fill_type='solid')
gold_fill = PatternFill(start_color=gold, end_color=gold, fill_type='solid')
white_fill = PatternFill(start_color=white, end_color=white, fill_type='solid')
green_fill = PatternFill(start_color=green, end_color=green, fill_type='solid')

# Tools dictionary. The integer values correspond to the column.
tpg_tools = {'Arctic Wolf Agent': 2,
             'SnapAgent': 3,
             'Umbrella Roaming Client': 4,
             'The Purple Guys Support Concierge':5,
             'Sophos Endpoint Agent':6,
             'Security Manager AV Defender': 7,
             }

# List of competing AV. Will update/add this list as needed.
competing_av = ['Cylance Protect',
                'Trend Micro',
                'ESET Endpoint Security',
                'webroot',
                'COMODO',
                'VIPRE',
                'Bitdefender',
                'Dell Protected Workspace',
                ]

# TODO add regexes for the competing AV above. TO make it more trustworthy.
#  So far, testing is going well though.
# REGEX block
# regex to find Client names
cust_regex = re.compile(r'''^.*(Customer: (.*?))Executed By:''')
# regex to find Client names v2
client_regex = re.compile(r'''^Scheduled Task (.*?) -''')
# regex to find job type and amp/script name
type_regex = re.compile(r'''^.*Type: (.*?) \[(.*?)\]''')
# regex to find device name
device_regex = re.compile(r'''^.*Device: (.*?)\[''')
# regex to find bde status output
bde_regex = re.compile(r'''(Conversion Status: )(.*?) (Percentage)''')
# regex to find TPM status
tpm_regex = re.compile(r'''oscpresent:(.*?)
                           oscactive:(.*?)
                           oscenabled:(.*?)
                           Result''', re.VERBOSE)
# regex to find zip files
zip_regex = re.compile(r"""^(.*?)(\.)(zip)$""")
# regex to find txt files
txt_regex = re.compile(r"""^(.*?)(\.)(txt)$""")

# iterate through all emails and process (Main block)
for msg in list(messages):
    # parse info from email body, organize into variables, and handle errors
    cust_mo = re.search(cust_regex, msg.Body)
    client_mo = re.search(client_regex, msg.Subject)
    type_mo = re.search(type_regex, msg.Body)
    device_mo = re.search(device_regex, msg.Body)

    # Parse Client name
    if cust_mo:
        client_name = cust_mo.group(2).strip()
        logging.debug(f'Client: {client_name}')
    else:
        with open(err_file, 'a') as file:
            file.write(f'\nNo Customer Detected. '
                       f'Skipping Email Subject: {msg.Subject}...')
        continue

    # Client name version 2. If set correctly, overwrite the above.
    if client_mo:
        if client_mo.group(1).strip() != 'Automation Policy':
            client_name = client_mo.group(1).strip()
            logging.debug(f'Client v2: {client_name}')

    # client folder management (now that we have client name)
    if os.path.exists(parent_f + client_name) is False:
        os.makedirs(parent_f + client_name)
    client_folder = parent_f + client_name + '\\'
    if os.path.exists(client_folder + 'temp') is False:
        os.makedirs(client_folder + 'temp')
    client_temp = client_folder + 'temp\\'
    client_err_file = client_folder + f'{client_name}_AMP_errors.txt'
    logging.debug(f'Client Folder location: {client_folder}')

    # Parse job type and name (job type is not really used yet,
    # but leaving it for future support if needed)
    if type_mo:
        job_type = type_mo.group(1).strip()
        logging.debug(f'Job type: {job_type}')
        job_name = type_mo.group(2).strip()
        logging.debug(f'Job name: {job_name}')
    else:
        with open(client_err_file, 'a') as file:
            file.write(f'\nNo Job Detected. '
                       f'Skipping Email Subject: {msg.Subject}...')
        continue

    # Parse device name
    if device_mo:
        device_name = device_mo.group(1).strip()
        logging.debug(f'Device name: {device_name}')
    else:
        with open(client_err_file, 'a') as file:
            file.write(f'\nNo Device Detected. '
                       f'Skipping Email Subject: {msg.Subject}...')
        continue

    # Added check to handle some occasional false "successes"
    if 'Task did not produce any output.' in msg.Body:
        with open(client_err_file, 'a') as file:
            file.write(f'\n{client_name}: '
                       f'{device_name} did not produce any output for '
                       f'{job_name}')
        continue

    # Added check for another possible error
    if 'This policy has encountered an exception' in msg.Body:
        with open(client_err_file, 'a') as file:
            file.write(f'\n{client_name}: '
                       f'{device_name} failed to run '
                       f'{job_name}')
        continue

    # While keeping track of file's parent company, job, read output contents,
    #  and update client_name spreadsheet with device and details
    wb = Workbook()
    wb_file = client_folder + f'{client_name}.xlsx'

    # Check if client xlsx exists, if not create, and prep
    #TODO put this is a little function. Start putting in spearate books.
    # Also, new book/support for pulling encrpytion keys during offboarding.
    if os.path.exists(wb_file) is False:
        wb.save(wb_file)
        wb = load_workbook(wb_file)
        encrypt_sheet = wb['Sheet']
        encrypt_sheet.title = 'Encryption'
        board_sheet = wb.create_sheet('On-Offboard')
        font_header = Font(size=12, bold=True)
        encrypt_headers = [('Device Name', 'TPM Present?', 'TPM Active?',
                   'TPM Enabled?', 'Encryption Status')]
        board_headers = [('Device Name', 'Arctic Wolf?', 'Blackpoint SNAP?',
                          'Umbrella?', 'Concierge?', 'Sophos?',
                          'AV Defender?', 'Competing AV?')]
        for sheet in wb.worksheets:
            for i in range(1, 9):
                col = get_column_letter(i)
                sheet.column_dimensions[col].width = 25
        for row in encrypt_headers:
            encrypt_sheet.append(row)
        for row in board_headers:
            board_sheet.append(row)
        for cell in encrypt_sheet['1:1']:
            cell.font = font_header
        for cell in board_sheet['1:1']:
            cell.font = font_header
        encrypt_sheet.freeze_panes = 'A2'
        board_sheet.freeze_panes = 'A2'
        wb.save(wb_file)

    # Load client workbook
    wb = load_workbook(wb_file)
    encrypt_sheet = wb['Encryption']
    board_sheet = wb['On-Offboard']

    # Handle TPM amp and populate spreadsheet
    if job_name == 'Windows TPM Monitoring':
        device_row = device_check(encrypt_sheet, device_name)
        tpm_mo = re.search(tpm_regex, msg.Body)
        if tpm_mo:
            tpm_present = tpm_mo.group(1).strip()
            tpm_active = tpm_mo.group(2).strip()
            tpm_enabled = tpm_mo.group(3).strip()
            logging.debug(f'tpm present?: {tpm_present}')
            logging.debug(f'tpm active?: {tpm_active}')
            logging.debug(f'tpm enabled?: {tpm_enabled}')
            if device_row == '':
                new_row = [(device_name, tpm_present,
                            tpm_active, tpm_enabled)]
                device_row = encrypt_sheet.max_row + 1
                for row in new_row:
                    encrypt_sheet.append(row)
            else:
                encrypt_sheet.cell(row=device_row, column=2).value = tpm_present
                encrypt_sheet.cell(row=device_row, column=3).value = tpm_active
                encrypt_sheet.cell(row=device_row, column=4).value = tpm_enabled

            # highlight any row with no TPM detected
            if tpm_present == 'No TPM Detected':
                for cell in encrypt_sheet[device_row]:
                    cell.fill = gold_fill

    # Handle encryption status check
    elif job_name == 'manage-bde -status':
        device_row = device_check(encrypt_sheet, device_name)
        bde_mo = re.search(bde_regex, msg.Body)
        if bde_mo:
            encrypt_status = bde_mo.group(2).strip()
            logging.debug(f'Encrypted status: {encrypt_status}')
            if device_row == '':
                new_row = [(device_name, '', '', '', encrypt_status)]
                for row in new_row:
                    encrypt_sheet.append(row)
            else:
                encrypt_sheet.cell(
                    row=device_row, column=5).value = encrypt_status

    # Handle Software inventory
    elif job_name == 'Simple Software Inventory':
        device_row = device_check(board_sheet, device_name)
        for attach in msg.Attachments:
            mo = re.search(zip_regex, str(attach))
            if mo:
                # If found, save attachment
                temp_file = client_temp + device_name + '_' + attach.FileName
                attach.SaveAsFile(temp_file)

                # Unzip file
                with ZipFile(temp_file) as zip_obj:
                    zip_obj.extractall(path=client_temp)

                # There should only be one, but iterate anyway
                files = os.listdir(client_temp)
                for f in files:
                    txt_mo = re.search(txt_regex, f)
                    if txt_mo:
                        # open and parse
                        with open(client_temp + f, 'r',
                                  encoding='utf-8') as out_file:
                            app_data = out_file.read()
                    else:
                        continue
                    if device_row == '':
                        device_row = board_sheet.max_row + 1
                    board_sheet.cell(row=device_row,
                                     column=1).value = device_name
                    # Look for TPG apps
                    for k, v in tpg_tools.items():
                        if k in app_data:
                            board_sheet.cell(row=device_row,
                                             column=v).value= 'Installed'
                            board_sheet.cell(row=device_row,
                                             column=v).fill = green_fill
                        else:
                            board_sheet.cell(row=device_row,
                                             column=v).value = 'Missing'
                            board_sheet.cell(row=device_row,
                                             column=v).fill = red_fill

                    # Look for competing AV, break loop when first is found.
                    # Others may be present,
                    # but we'll be working on this PC anyway if one is found
                    for apps in competing_av:
                        if apps in app_data:
                            board_sheet.cell(row=device_row,
                                             column=8).value = apps
                            board_sheet.cell(row=device_row,
                                             column=8).fill = gold_fill
                            break
                        else:
                            board_sheet.cell(row=device_row,
                                             column=8).value = 'None Found'
        # Delete the temp files when finished.
        temp_files = os.listdir(client_temp)
        for f in temp_files:
            os.remove(client_temp + f)

    # Handle anything else right now
    else:
        with open(client_err_file, 'a') as file:
            file.write(f'\nNo support added for {job_name} yet. Sorry. '
                       f'Skipping {client_name}: '
                       f'{device_name}: {job_name}...')
        logging.debug('End of msg process\n')
        continue

    # Save the spreadsheet
    wb.save(wb_file)

    # Move the processed message to the processed folder.
    msg.Move(outbox)
    logging.debug('End of msg process\n')

logging.debug('End of program')
